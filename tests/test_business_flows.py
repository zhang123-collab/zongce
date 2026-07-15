import asyncio
import importlib.util
import io
import json
import os
import tempfile
import unittest


RUNTIME_AVAILABLE = all(
    importlib.util.find_spec(name)
    for name in ("fastapi", "sqlalchemy", "jose", "passlib", "multipart", "openpyxl", "reportlab")
)


@unittest.skipUnless(RUNTIME_AVAILABLE, "项目运行依赖尚未安装")
class BusinessFlowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory(prefix="zongce-tests-")
        db_path = os.path.join(cls.temp_dir.name, "test.db").replace("\\", "/")
        os.environ["ZONGCE_DATABASE_URL"] = f"sqlite:///{db_path}"
        os.environ["ZONGCE_UPLOAD_DIR"] = os.path.join(cls.temp_dir.name, "uploads")
        os.environ["ZONGCE_SECRET_KEY"] = "test-only-secret"
        global app_module
        import app as app_module

    @classmethod
    def tearDownClass(cls):
        app_module.engine.dispose()
        cls.temp_dir.cleanup()

    def setUp(self):
        self.db = app_module.SessionLocal()

    def tearDown(self):
        self.db.rollback()
        self.db.close()

    def student_for_grade(self, grade):
        return (
            self.db.query(app_module.StudentProfile)
            .filter(app_module.StudentProfile.grade == grade)
            .first()
        )

    def first_rule(self):
        return self.db.query(app_module.ScoreRule).filter(app_module.ScoreRule.is_active == True).first()

    def create_accounting_student(self, suffix):
        user = app_module.User(
            username=f"accounting_{suffix}",
            password=app_module.get_password_hash("123456"),
            role=app_module.ROLE_STUDENT,
            real_name=f"核算测试{suffix}",
        )
        self.db.add(user)
        self.db.flush()
        student = app_module.StudentProfile(
            user_id=user.id,
            student_id=f"2099{suffix:08d}",
            class_name="核算测试班",
            major="人工智能",
            grade="2099",
            moral_score=10,
            academic_score=50,
        )
        self.db.add(student)
        self.db.commit()
        return student

    def make_batch_workbook(self, student, rule, declared_student_id=None):
        from openpyxl import Workbook
        workbook = Workbook()
        summary = workbook.active
        summary.title = "学生汇总"
        summary.append(["姓名", "学号"])
        summary.append([student.user.real_name, declared_student_id or student.student_id])
        details = workbook.create_sheet("申报明细")
        details.append(["申报明细（学生逐项填写，系统优先读取本页）"] * 16)
        details.append([
            "明细序号", "模块", "汇总类别", "项目/活动名称", "规则等级/获奖结果", "本人身份/排名",
            "学生申报分", "证据编号", "证据类型", "公用材料页码", "个人材料页码", "证据说明/关键词",
            "发生/获奖/任职时间", "姓名核验", "时间核验", "备注",
        ])
        details.append([
            1, rule.category, rule.sub_category, rule.item_name, "校级", "第1名", 1,
            "E01", "个人证明", "", "P1", "回归测试证据", "2026-01-02", "待核验", "待核验", "测试",
        ])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def test_split_routers_are_registered(self):
        def collect_paths(routes):
            result = set()
            for route in routes:
                if hasattr(route, "path"):
                    result.add(route.path)
                if hasattr(route, "routes"):
                    result.update(collect_paths(route.routes))
            return result
        paths = collect_paths(app_module.app.routes)
        for path in (
            "/", "/api/login", "/api/student/profile",
            "/api/admin/student/import", "/api/application/create", "/api/file/{file_id}",
            "/api/audit/pass", "/api/admin/rule/list", "/api/admin/class/list", "/static",
            "/api/accounting/me", "/api/accounting/list", "/api/admin/accounting/export",
            "/api/batch/upload", "/api/batch/my", "/api/batch/{batch_id}/confirm", "/api/batch/template",
            "/api/ai/status", "/api/ai/batch/{batch_id}/analyze", "/api/ai/job/{job_id}",
            "/api/ai/application/{application_id}/analyze", "/api/ai/application/{application_id}/latest",
            "/api/announcement/list", "/api/admin/announcement", "/api/objection",
            "/api/admin/objection/{objection_id}/handle", "/api/evidence/duplicates",
            "/api/report/student/{student_id}.pdf", "/api/admin/accounting/recalculate-ranks",
        ):
            self.assertIn(path, paths)
        tables = set(app_module.inspect(app_module.engine).get_table_names())
        self.assertTrue({
            "score_deduction", "score_finalization", "operation_log",
            "submission_batch", "submission_batch_item", "submission_batch_material",
            "ai_analysis_job", "ai_item_suggestion",
            "announcement", "score_objection", "application_ai_analysis",
        }.issubset(tables))

    def create_draft(self, student_profile):
        request = app_module.ApplicationCreate(
            rule_id=self.first_rule().id,
            project_name="回归测试申请",
            team_rank=1,
            team_total=1,
        )
        result = app_module.create_application(request, student_profile.user, self.db)
        self.assertEqual(result["code"], 200)
        return result["data"]["id"]

    def test_submit_requires_uploaded_evidence(self):
        student = self.student_for_grade("2023")
        app_id = self.create_draft(student)
        rejected = app_module.submit_application(app_id, student.user, self.db)
        self.assertEqual(rejected["code"], 400)

        upload = app_module.UploadFile(filename="证明.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        uploaded = asyncio.run(app_module.upload_file(app_id, upload, student.user, self.db))
        self.assertEqual(uploaded["code"], 200)
        submitted = app_module.submit_application(app_id, student.user, self.db)
        self.assertEqual(submitted["code"], 200)

    def test_upload_rejects_content_disguised_as_pdf(self):
        student = self.student_for_grade("2023")
        app_id = self.create_draft(student)
        upload = app_module.UploadFile(filename="伪装材料.pdf", file=io.BytesIO(b"not a real pdf"))
        result = asyncio.run(app_module.upload_file(app_id, upload, student.user, self.db))
        self.assertEqual(result["code"], 400)
        self.assertIn("扩展名", result["message"])
        evidence_count = self.db.query(app_module.EvidenceFile).filter_by(application_id=app_id).count()
        self.assertEqual(evidence_count, 0)

    def test_request_validation_error_uses_api_response_shape(self):
        from fastapi.exceptions import RequestValidationError
        exc = RequestValidationError([
            {
                "type": "missing",
                "loc": ("body", "username"),
                "msg": "Field required",
                "input": {},
            }
        ])
        response = asyncio.run(app_module.request_validation_exception_handler(None, exc))
        body = json.loads(response.body)
        self.assertEqual(response.status_code, 422)
        self.assertEqual(body["code"], 422)
        self.assertEqual(body["data"]["errors"][0]["field"], "username")

    def test_direct_create_and_submit_is_rejected(self):
        student = self.student_for_grade("2023")
        request = app_module.ApplicationCreate(
            rule_id=self.first_rule().id,
            project_name="不能绕过材料",
            submit_now=True,
        )
        result = app_module.create_application(request, student.user, self.db)
        self.assertEqual(result["code"], 400)

    def test_rejected_application_returns_to_draft_when_edited(self):
        student = self.student_for_grade("2023")
        app_id = self.create_draft(student)
        record = self.db.query(app_module.ScoreApplication).filter_by(id=app_id).first()
        record.status = app_module.STATUS_REJECTED
        record.reject_reason = "需要修改"
        self.db.commit()

        result = app_module.update_application(
            app_id,
            app_module.ApplicationUpdate(project_name="修改后的申请"),
            student.user,
            self.db,
        )
        self.assertEqual(result["code"], 200)
        self.db.refresh(record)
        self.assertEqual(record.status, app_module.STATUS_DRAFT)
        self.assertEqual(record.reject_reason, "")

    def test_teacher_cannot_audit_another_class(self):
        student = self.student_for_grade("2022")
        self.assertIsNotNone(student)
        app_id = self.create_draft(student)
        record = self.db.query(app_module.ScoreApplication).filter_by(id=app_id).first()
        record.status = app_module.STATUS_PENDING
        self.db.commit()
        teacher = self.db.query(app_module.User).filter_by(username="teacher01").first()

        result = app_module.audit_pass(
            app_module.AuditPassReq(application_id=app_id),
            teacher,
            self.db,
        )
        self.assertEqual(result["code"], 403)

    def test_login_is_locked_after_repeated_failures(self):
        from zongce.routers.members import _do_login
        user = self.student_for_grade("2022").user
        for _ in range(app_module.MAX_LOGIN_ATTEMPTS):
            result = _do_login(user.username, "错误密码", self.db)
        self.assertEqual(result["code"], 400)
        locked = _do_login(user.username, "123456", self.db)
        self.assertEqual(locked["code"], 423)
        user.failed_login_count = 0
        user.locked_until = None
        self.db.commit()

    def test_admin_can_import_student_workbook(self):
        from openpyxl import Workbook
        from zongce.routers.members import admin_student_import
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["学号", "姓名", "班级", "专业", "年级", "思品分", "学业成绩"])
        sheet.append(["202699999999", "导入测试", "2026级1班", "计算机类", "2026", 80, 90])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = app_module.UploadFile(filename="students.xlsx", file=buffer)
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        result = asyncio.run(admin_student_import(upload, admin, self.db))
        self.assertEqual(result["code"], 200)
        self.assertEqual(result["data"]["success"], 1)

    def pending_application(self):
        student = self.student_for_grade("2023")
        app_id = self.create_draft(student)
        upload = app_module.UploadFile(filename="audit.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        asyncio.run(app_module.upload_file(app_id, upload, student.user, self.db))
        app_module.submit_application(app_id, student.user, self.db)
        return app_id

    def test_audit_can_explicitly_modify_score_to_zero(self):
        from zongce.routers.audits import AuditModifyReq, audit_modify
        app_id = self.pending_application()
        teacher = self.db.query(app_module.User).filter_by(username="teacher01").first()
        result = audit_modify(
            AuditModifyReq(application_id=app_id, modified_score=0, opinion="核验后按零分处理"),
            teacher,
            self.db,
        )
        self.assertEqual(result["code"], 200)
        record = self.db.query(app_module.ScoreApplication).filter_by(id=app_id).first()
        self.assertEqual(record.status, app_module.STATUS_PASSED)
        self.assertEqual(record.final_score, 0)

    def test_audit_opinion_length_is_limited(self):
        from zongce.routers.audits import AuditPassReq, audit_pass
        app_id = self.pending_application()
        teacher = self.db.query(app_module.User).filter_by(username="teacher01").first()
        result = audit_pass(AuditPassReq(application_id=app_id, opinion="a" * 501), teacher, self.db)
        self.assertEqual(result["code"], 400)

    def test_audit_history_tolerates_unknown_legacy_result(self):
        from zongce.routers.audits import audit_history
        student = self.student_for_grade("2023")
        app_id = self.create_draft(student)
        self.db.add(app_module.AuditRecord(
            application_id=app_id,
            auditor_id=student.user.id,
            result=99,
            opinion="旧数据兼容测试",
        ))
        self.db.commit()
        result = audit_history(app_id, student.user, self.db)
        self.assertEqual(result["code"], 200)
        self.assertEqual(result["data"][0]["resultName"], "未知")

    def test_admin_cannot_disable_current_account(self):
        from zongce.routers.admin import AdminEditUserReq, admin_edit_user
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        result = admin_edit_user(AdminEditUserReq(user_id=admin.id, is_active=False), admin, self.db)
        self.assertEqual(result["code"], 400)

    def test_class_name_rejects_script_characters(self):
        from zongce.routers.admin import ClassCreateReq, admin_class_create
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        result = admin_class_create(ClassCreateReq(class_name="<script>"), admin, self.db)
        self.assertEqual(result["code"], 400)

    def test_accounting_aggregates_caps_and_soft_deleted_deductions(self):
        from zongce.accounting_service import calculate_student_accounting
        from zongce.routers.accounting import DeductionCreateReq, accounting_add_deduction, accounting_delete_deduction
        student = self.create_accounting_student(1)
        rule = app_module.ScoreRule(
            category="学术创新成果",
            sub_category="核算测试项目",
            item_name="核算测试项目",
            base_score=1,
            max_score=6,
            policy=app_module.POLICY_SUM,
            is_active=True,
        )
        self.db.add(rule)
        self.db.flush()
        self.db.add_all([
            app_module.ScoreApplication(student_id=student.id, rule_id=rule.id, project_name="项目A", status=app_module.STATUS_PASSED, final_score=5),
            app_module.ScoreApplication(student_id=student.id, rule_id=rule.id, project_name="项目B", status=app_module.STATUS_PASSED, final_score=4),
        ])
        self.db.commit()
        summary = calculate_student_accounting(self.db, student)
        self.assertEqual(summary["innovationScore"], 6)
        self.assertEqual(summary["totalScore"], 66)

        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        added = accounting_add_deduction(
            DeductionCreateReq(student_id=student.student_id, deduction_score=3, reason="核算测试扣分"),
            admin,
            self.db,
        )
        self.assertEqual(added["code"], 200)
        self.assertEqual(added["data"]["totalScore"], 63)
        deduction_id = added["data"]["deductions"][0]["id"]
        removed = accounting_delete_deduction(deduction_id, admin, self.db)
        self.assertEqual(removed["code"], 200)
        self.assertEqual(removed["data"]["totalScore"], 66)
        deduction = self.db.query(app_module.ScoreDeduction).filter_by(id=deduction_id).first()
        self.assertFalse(deduction.is_active)

    def test_finalization_locks_application_and_can_be_reopened_with_reason(self):
        from zongce.routers.accounting import ReopenReq, accounting_finalize, accounting_reopen
        student = self.create_accounting_student(2)
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        finalized = accounting_finalize(student.student_id, admin, self.db)
        self.assertEqual(finalized["code"], 200)
        self.assertTrue(finalized["data"]["isFinalized"])

        create_result = app_module.create_application(
            app_module.ApplicationCreate(rule_id=self.first_rule().id, project_name="终审后不能创建"),
            student.user,
            self.db,
        )
        self.assertEqual(create_result["code"], 409)
        reopened = accounting_reopen(ReopenReq(student_id=student.student_id, reason="复核材料后重新开放"), admin, self.db)
        self.assertEqual(reopened["code"], 200)
        self.assertFalse(reopened["data"]["isFinalized"])

    def test_teacher_accounting_list_is_scoped_to_managed_class(self):
        from zongce.routers.accounting import accounting_list
        teacher = self.db.query(app_module.User).filter_by(username="teacher01").first()
        result = accounting_list("", "", teacher, self.db)
        self.assertEqual(result["code"], 200)
        self.assertTrue(all(item["className"] == teacher.managed_class for item in result["data"]["list"]))

    def test_accounting_export_returns_xlsx_stream(self):
        from zongce.routers.accounting import accounting_export
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        response = accounting_export(admin, self.db)
        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response.headers["content-disposition"])

    def test_batch_upload_parses_excel_and_confirms_into_draft_applications(self):
        from zongce.routers.batches import batch_confirm, batch_upload
        student = self.create_accounting_student(3)
        rule = self.first_rule()
        excel = app_module.UploadFile(filename="申报.xlsx", file=self.make_batch_workbook(student, rule))
        material = app_module.UploadFile(filename="证明.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        uploaded = asyncio.run(batch_upload(excel, [material], student.user, self.db))
        self.assertEqual(uploaded["code"], 200)
        self.assertEqual(uploaded["data"]["status"], "parsed")
        self.assertEqual(uploaded["data"]["validCount"], 1)
        batch_id = uploaded["data"]["id"]

        confirmed = batch_confirm(batch_id, student.user, self.db)
        self.assertEqual(confirmed["code"], 200)
        application_id = confirmed["data"]["applicationIds"][0]
        application = self.db.query(app_module.ScoreApplication).filter_by(id=application_id).first()
        self.assertEqual(application.status, app_module.STATUS_DRAFT)
        self.assertEqual(len(application.evidence_files), 1)
        repeated = batch_confirm(batch_id, student.user, self.db)
        self.assertEqual(repeated["code"], 400)

    def test_batch_identity_mismatch_is_reported_by_excel_row(self):
        from zongce.routers.batches import batch_upload
        student = self.create_accounting_student(4)
        rule = self.first_rule()
        excel = app_module.UploadFile(
            filename="错误学号.xlsx",
            file=self.make_batch_workbook(student, rule, declared_student_id="200000000000"),
        )
        material = app_module.UploadFile(filename="证明.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        uploaded = asyncio.run(batch_upload(excel, [material], student.user, self.db))
        self.assertEqual(uploaded["code"], 200)
        self.assertEqual(uploaded["data"]["status"], "needs_correction")
        self.assertIn("学号", uploaded["data"]["items"][0]["errorMessage"])

    def test_batch_rejects_fake_excel_content(self):
        from zongce.routers.batches import batch_upload
        student = self.create_accounting_student(5)
        excel = app_module.UploadFile(filename="伪装.xlsx", file=io.BytesIO(b"PK\x03\x04not-a-real-xlsx"))
        material = app_module.UploadFile(filename="证明.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        result = asyncio.run(batch_upload(excel, [material], student.user, self.db))
        self.assertEqual(result["code"], 400)
        self.assertIn("XLSX", result["message"])

    def test_ai_redaction_removes_identity_and_local_paths(self):
        from zongce.ai.redaction import ensure_private, redact_text
        raw = (
            "核算测试6，学号209900000006，手机 +86 138-1234-5678，"
            "邮箱student@example.com，文件C:\\Users\\student\\秘密证明.pdf，编号1234-5678-9012，"
            "名单：张三 202311081001，202311081002 李四，日期2026-01-02"
        )
        result = redact_text(raw, "核算测试6", "209900000006", ["秘密证明.pdf"])
        for secret in ("核算测试6", "209900000006", "138-1234-5678", "student@example.com", "秘密证明.pdf", "C:\\Users", "张三", "李四", "202311081001", "202311081002"):
            self.assertNotIn(secret, result)
        self.assertIn("[姓名]", result)
        self.assertIn("[学号]", result)
        self.assertIn("[手机号]", result)
        self.assertIn("2026-01-02", result)
        ensure_private(result, ["核算测试6", "209900000006", "秘密证明.pdf"])
        with self.assertRaises(ValueError):
            ensure_private("仍残留学号202311081099")

    def test_ai_job_saves_advice_without_changing_application_scores(self):
        from unittest.mock import patch
        from zongce.ai.task_service import run_analysis_job
        from zongce.routers.batches import batch_upload

        student = self.create_accounting_student(6)
        rule = self.first_rule()
        excel = app_module.UploadFile(filename="申报.xlsx", file=self.make_batch_workbook(student, rule))
        material = app_module.UploadFile(filename="个人秘密证明.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        uploaded = asyncio.run(batch_upload(excel, [material], student.user, self.db))
        batch_id = uploaded["data"]["id"]
        item_id = uploaded["data"]["items"][0]["id"]
        job = app_module.AiAnalysisJob(
            batch_id=batch_id,
            student_id=student.id,
            status="pending",
            item_count=1,
        )
        self.db.add(job)
        self.db.commit()
        job_id = job.id
        with patch("zongce.ai.task_service.call_deepseek_batch", return_value=[{
            "item_id": item_id,
            "verification_status": "匹配",
            "suggested_score": 2,
            "selected_rule_id": rule.id,
            "reason": "材料与申报项目相符，建议人工复核",
        }]):
            run_analysis_job(job_id)
        self.db.expire_all()
        saved_job = self.db.query(app_module.AiAnalysisJob).filter_by(id=job_id).first()
        suggestion = self.db.query(app_module.AiItemSuggestion).filter_by(job_id=job_id).first()
        self.assertEqual(saved_job.status, "completed")
        self.assertEqual(suggestion.verification_status, "匹配")
        self.assertEqual(suggestion.suggested_score, 2)
        self.assertEqual(
            self.db.query(app_module.ScoreApplication).filter_by(student_id=student.id).count(),
            0,
        )
        run_analysis_job(job_id)
        self.assertEqual(
            self.db.query(app_module.AiItemSuggestion).filter_by(job_id=job_id).count(),
            1,
        )

    def test_ai_job_permissions_follow_student_and_managed_class(self):
        from zongce.routers.ai import _can_access_job
        owner = self.create_accounting_student(7)
        other = self.create_accounting_student(8)
        batch = app_module.SubmissionBatch(
            student_id=owner.id,
            assessment_year="2025-2026",
            status="parsed",
            excel_name="test.xlsx",
            excel_path="test.xlsx",
        )
        self.db.add(batch)
        self.db.flush()
        job = app_module.AiAnalysisJob(batch_id=batch.id, student_id=owner.id)
        self.db.add(job)
        self.db.commit()
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        self.assertTrue(_can_access_job(self.db, owner.user, job))
        self.assertFalse(_can_access_job(self.db, other.user, job))
        self.assertTrue(_can_access_job(self.db, admin, job))

    def test_dense_class_and_grade_ranking(self):
        from zongce.accounting_service import calculate_ranked_accounting
        first = self.create_accounting_student(20)
        second = self.create_accounting_student(21)
        third = self.create_accounting_student(22)
        first.academic_score = 88
        second.academic_score = 88
        third.academic_score = 70
        self.db.commit()
        results = calculate_ranked_accounting(self.db, [first, second, third])
        by_id = {item["studentId"]: item for item in results}
        self.assertEqual(by_id[first.student_id]["classRank"], by_id[second.student_id]["classRank"])
        self.assertGreater(by_id[third.student_id]["classRank"], by_id[first.student_id]["classRank"])
        self.assertEqual(by_id[first.student_id]["gradeRank"], by_id[second.student_id]["gradeRank"])

    def test_cached_ranking_only_recalculates_requested_student(self):
        from unittest.mock import patch
        import zongce.accounting_service as service
        student = self.create_accounting_student(27)
        service.calculate_ranked_accounting(self.db, persist=True)
        with patch(
            "zongce.accounting_service.calculate_student_accounting",
            wraps=service.calculate_student_accounting,
        ) as calculate:
            result = service.calculate_ranked_accounting(self.db, [student])
        self.assertEqual(len(result), 1)
        self.assertEqual(calculate.call_count, 1)

    def test_publication_objection_workflow(self):
        from datetime import datetime, timedelta
        from zongce.routers.publications import (
            ObjectionHandleReq, ObjectionReq, create_objection, handle_objection,
        )
        student = self.create_accounting_student(23)
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        announcement = app_module.Announcement(
            title="回归测试公示",
            scope_type="class",
            scope_value=student.class_name,
            starts_at=datetime.now() - timedelta(hours=1),
            ends_at=datetime.now() + timedelta(days=2),
            created_by=admin.id,
        )
        self.db.add(announcement)
        self.db.commit()
        created = create_objection(ObjectionReq(
            announcement_id=announcement.id,
            score_item="综合测评总分",
            objection_type="分数计算错误",
            description="请复核加分汇总。",
        ), student.user, self.db)
        self.assertEqual(created["code"], 200)
        duplicate = create_objection(ObjectionReq(
            announcement_id=announcement.id,
            score_item="综合测评总分",
            objection_type="分数计算错误",
            description="重复提交。",
        ), student.user, self.db)
        self.assertEqual(duplicate["code"], 409)
        handled = handle_objection(
            created["data"]["id"],
            ObjectionHandleReq(resolution_status="accepted", resolution="已复核并进入修正流程。"),
            admin,
            self.db,
        )
        self.assertEqual(handled["data"]["status"], "accepted")

    def test_duplicate_material_hash_and_report_pdf(self):
        from zongce.accounting_service import calculate_ranked_accounting
        from zongce.report_service import build_student_report
        from zongce.routers.applications import duplicate_evidence_groups
        first = self.create_accounting_student(24)
        second = self.create_accounting_student(25)
        first_app = self.create_draft(first)
        second_app = self.create_draft(second)
        content = b"%PDF-1.4\nregression duplicate\n%%EOF"
        first_upload = asyncio.run(app_module.upload_file(
            first_app, app_module.UploadFile(filename="甲.pdf", file=io.BytesIO(content)), first.user, self.db,
        ))
        second_upload = asyncio.run(app_module.upload_file(
            second_app, app_module.UploadFile(filename="乙.pdf", file=io.BytesIO(content)), second.user, self.db,
        ))
        self.assertFalse(first_upload["data"]["duplicateWarning"])
        self.assertTrue(second_upload["data"]["duplicateWarning"])
        admin = self.db.query(app_module.User).filter_by(username="admin").first()
        groups = duplicate_evidence_groups(admin, self.db)
        self.assertTrue(any(group["count"] >= 2 for group in groups["data"]["groups"]))
        result = calculate_ranked_accounting(self.db, [first])[0]
        pdf = build_student_report(result).read()
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 1000)

    def test_single_application_ai_is_advisory(self):
        from unittest.mock import patch
        from zongce.ai.application_service import run_application_analysis
        student = self.create_accounting_student(26)
        application_id = self.create_draft(student)
        upload = app_module.UploadFile(filename="单条申请材料.pdf", file=io.BytesIO(b"%PDF-1.4\n%%EOF"))
        asyncio.run(app_module.upload_file(application_id, upload, student.user, self.db))
        application = self.db.query(app_module.ScoreApplication).filter_by(id=application_id).first()
        original_score = application.system_calculated_score
        analysis = app_module.ApplicationAiAnalysis(
            application_id=application.id,
            student_id=student.id,
            status="pending",
        )
        self.db.add(analysis)
        self.db.commit()
        with patch("zongce.ai.application_service.call_deepseek_batch", return_value=[{
            "item_id": application.id,
            "verification_status": "匹配",
            "suggested_score": original_score,
            "selected_rule_id": application.rule_id,
            "reason": "材料与申报信息相符，建议人工复核。",
        }]):
            run_application_analysis(analysis.id)
        self.db.expire_all()
        saved = self.db.query(app_module.ApplicationAiAnalysis).filter_by(id=analysis.id).first()
        application = self.db.query(app_module.ScoreApplication).filter_by(id=application_id).first()
        self.assertEqual(saved.status, "completed")
        self.assertEqual(saved.verification_status, "匹配")
        self.assertEqual(application.system_calculated_score, original_score)
        self.assertEqual(application.status, app_module.STATUS_DRAFT)


if __name__ == "__main__":
    unittest.main()
