from datetime import date
from pathlib import Path
from uuid import uuid4
import json
import io
import os
import re
import shutil
import zipfile

from openpyxl import load_workbook

from zongce.core import *
from zongce.accounting_service import ASSESSMENT_YEAR, is_student_finalized, log_operation


BATCH_PARSED = "parsed"
BATCH_NEEDS_CORRECTION = "needs_correction"
BATCH_CONFIRMED = "confirmed"
BATCH_MAX_ITEMS = 100
BATCH_MAX_MATERIALS = 5
EXCEL_MAX_BYTES = 5 * 1024 * 1024
MATERIAL_MAX_BYTES = 10 * 1024 * 1024
TOTAL_MATERIAL_MAX_BYTES = 30 * 1024 * 1024

DETAIL_HEADERS = {
    "模块", "汇总类别", "项目/活动名称", "规则等级/获奖结果", "本人身份/排名",
    "学生申报分", "证据编号", "证据类型", "公用材料页码", "个人材料页码",
    "证据说明/关键词", "发生/获奖/任职时间", "备注",
}


def safe_original_name(filename: str) -> str:
    name = os.path.basename((filename or "").replace("\\", "/"))
    if not name or len(name) > 255:
        raise ValueError("文件名不能为空且不能超过255个字符")
    return name


def validate_excel_content(filename: str, content: bytes) -> str:
    name = safe_original_name(filename)
    if not name.lower().endswith(".xlsx"):
        raise ValueError("申报表必须是 .xlsx 文件")
    if not content or len(content) > EXCEL_MAX_BYTES:
        raise ValueError("申报表不能为空且不能超过5MB")
    if not content.startswith(b"PK\x03\x04"):
        raise ValueError("申报表内容不是有效的XLSX文件")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > 1000 or sum(entry.file_size for entry in entries) > 20 * 1024 * 1024:
                raise ValueError("申报表解压内容过大或文件项过多")
            if any(".." in Path(entry.filename).parts or entry.filename.startswith(("/", "\\")) for entry in entries):
                raise ValueError("申报表内部路径不安全")
            if "[Content_Types].xml" not in archive.namelist():
                raise ValueError("申报表缺少XLSX必要结构")
    except zipfile.BadZipFile:
        raise ValueError("申报表内容不是有效的XLSX文件")
    return name


def validate_material_content(filename: str, content: bytes) -> tuple[str, str, str]:
    name = safe_original_name(filename)
    ext = Path(name).suffix.lower()
    signatures = {
        ".pdf": (b"%PDF", "application/pdf"),
        ".png": (b"\x89PNG\r\n\x1a\n", "image/png"),
        ".jpg": (b"\xff\xd8\xff", "image/jpeg"),
        ".jpeg": (b"\xff\xd8\xff", "image/jpeg"),
    }
    if ext not in signatures:
        raise ValueError(f"材料 {name} 仅支持PDF、JPG、PNG")
    if not content or len(content) > MATERIAL_MAX_BYTES:
        raise ValueError(f"材料 {name} 不能为空且不能超过10MB")
    signature, media_type = signatures[ext]
    if not content.startswith(signature):
        raise ValueError(f"材料 {name} 的内容与扩展名不匹配")
    return name, ext, media_type


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _score(value) -> tuple[float, str]:
    if value in (None, ""):
        return 0.0, ""
    try:
        result = float(value)
    except (TypeError, ValueError):
        return 0.0, "学生申报分不是有效数字"
    if not 0 <= result <= 100:
        return 0.0, "学生申报分应在0-100之间"
    return result, ""


def _date(value) -> tuple[str, str]:
    text_value = _text(value)
    if not text_value:
        return "", ""
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text_value, pattern).strftime("%Y-%m-%d"), ""
        except ValueError:
            continue
    return "", "日期格式应为YYYY-MM-DD"


def _match_rule(db: Session, module: str, category: str, project_name: str):
    rules = db.query(ScoreRule).filter(ScoreRule.is_active == True).all()
    category_text = (category or "").strip()
    module_text = (module or "").strip()
    name_text = (project_name or "").strip()
    exact = [rule for rule in rules if category_text and rule.sub_category == category_text]
    if exact:
        return exact[0]
    exact = [rule for rule in rules if category_text and rule.category == category_text and (rule.item_name in name_text or name_text in rule.item_name)]
    if exact:
        return exact[0]
    candidates = []
    haystack = f"{module_text} {category_text} {name_text}".lower()
    for rule in rules:
        score = 0
        for value in (rule.item_name, rule.sub_category, rule.category):
            needle = (value or "").strip().lower()
            if needle and (needle in haystack or name_text.lower() in needle):
                score += len(needle)
        if score:
            candidates.append((score, rule.id, rule))
    return max(candidates, default=(0, 0, None))[2]


def _summary_identity(workbook) -> tuple[str, str]:
    if "学生汇总" not in workbook.sheetnames:
        return "", ""
    sheet = workbook["学生汇总"]
    headers = [_text(cell.value) for cell in sheet[1]]
    values = [_text(cell.value) for cell in sheet[2]]
    data = dict(zip(headers, values))
    return data.get("姓名", ""), data.get("学号", "")


def parse_batch_excel(db: Session, excel_path: str, student: StudentProfile) -> list[dict]:
    # 文件已在上传阶段限制压缩体积和解压规模；非只读模式会在加载完成后关闭ZIP句柄，
    # 避免Windows上批次文件长期被占用而无法清理或替换。
    workbook = load_workbook(excel_path, read_only=False, data_only=True)
    declared_name, declared_student_id = _summary_identity(workbook)
    identity_error = ""
    if not declared_student_id:
        identity_error = "Excel学生汇总中的学号不能为空"
    elif declared_student_id != student.student_id:
        identity_error = f"Excel学号 {declared_student_id} 与当前登录学号不一致"
    if not declared_name:
        identity_error = (identity_error + "；" if identity_error else "") + "Excel学生汇总中的姓名不能为空"
    elif student.user and declared_name != student.user.real_name:
        identity_error = (identity_error + "；" if identity_error else "") + "Excel姓名与当前登录学生不一致"
    if "申报明细" not in workbook.sheetnames:
        raise ValueError("Excel缺少“申报明细”工作表")
    sheet = workbook["申报明细"]
    header_row = None
    headers = []
    for row_number in range(1, min(sheet.max_row, 10) + 1):
        values = [_text(cell.value) for cell in sheet[row_number]]
        if "项目/活动名称" in values:
            header_row = row_number
            headers = values
            break
    if not header_row:
        raise ValueError("未找到申报明细标准表头")
    missing = sorted(DETAIL_HEADERS - set(headers))
    if missing:
        raise ValueError("申报明细缺少列：" + "、".join(missing))
    result = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        data = dict(zip(headers, row))
        project_name = _text(data.get("项目/活动名称"))
        module = _text(data.get("模块"))
        category = _text(data.get("汇总类别"))
        if not project_name and not module and not category:
            continue
        errors = []
        if identity_error:
            errors.append(identity_error)
        if not project_name:
            errors.append("项目/活动名称不能为空")
        if len(project_name) > 200:
            errors.append("项目/活动名称不能超过200字")
        if not module:
            errors.append("模块不能为空")
        if not category:
            errors.append("汇总类别不能为空")
        declared_score, score_error = _score(data.get("学生申报分"))
        project_date, date_error = _date(data.get("发生/获奖/任职时间"))
        if score_error:
            errors.append(score_error)
        if date_error:
            errors.append(date_error)
        rule = _match_rule(db, module, category, project_name) if project_name else None
        if not rule:
            errors.append("未匹配到启用的加分规则，请检查类别和项目名称")
        evidence_parts = [
            _text(data.get("证据编号")), _text(data.get("证据类型")),
            _text(data.get("公用材料页码")), _text(data.get("个人材料页码")),
        ]
        description = _text(data.get("证据说明/关键词"))
        remark = _text(data.get("备注"))
        if len(description) > 500:
            errors.append("证据说明不能超过500字")
        if len(remark) > 200:
            errors.append("备注不能超过200字")
        result.append({
            "row_number": row_number,
            "rule_id": rule.id if rule else None,
            "project_name": project_name[:200],
            "project_level": _text(data.get("规则等级/获奖结果"))[:50],
            "role_rank": _text(data.get("本人身份/排名"))[:100],
            "declared_score": declared_score,
            "project_date": project_date,
            "description": description[:500],
            "evidence_ref": "；".join(part for part in evidence_parts if part)[:500],
            "remark": remark[:200],
            "status": "error" if errors else "valid",
            "error_message": "；".join(errors),
        })
        if len(result) > BATCH_MAX_ITEMS:
            raise ValueError(f"单次申报不能超过{BATCH_MAX_ITEMS}条明细")
    if not result:
        raise ValueError("申报明细中没有可读取的项目")
    return result


def calculate_rule_score(rule: ScoreRule, role_rank: str) -> float:
    coefficient = 1.0
    if rule.rank_coefficient:
        try:
            mapping = json.loads(rule.rank_coefficient)
            rank_match = re.search(r"\d+", role_rank or "")
            rank = rank_match.group(0) if rank_match else "default"
            coefficient = float(mapping.get(rank, mapping.get("default", 1.0)))
        except (TypeError, ValueError, json.JSONDecodeError):
            coefficient = 1.0
    return round(float(rule.base_score) * coefficient, 2)


def confirm_batch(db: Session, batch: SubmissionBatch, student: StudentProfile, actor: User) -> list[int]:
    if batch.student_id != student.id:
        raise ValueError("只能确认自己的批次")
    if batch.status == BATCH_CONFIRMED:
        raise ValueError("该批次已经确认")
    if batch.status != BATCH_PARSED or batch.error_count:
        raise ValueError("批次仍有解析错误，请修正Excel后重新上传")
    if is_student_finalized(db, student.id, batch.assessment_year):
        raise ValueError("综合测评结果已经终审，不能确认新批次")
    items = db.query(SubmissionBatchItem).filter(SubmissionBatchItem.batch_id == batch.id).order_by(SubmissionBatchItem.row_number).all()
    materials = db.query(SubmissionBatchMaterial).filter(
        SubmissionBatchMaterial.batch_id == batch.id,
        SubmissionBatchMaterial.is_active == True,
    ).all()
    if not materials:
        raise ValueError("请至少上传一份证明材料")
    application_ids = []
    for item in items:
        rule = db.query(ScoreRule).filter(ScoreRule.id == item.rule_id, ScoreRule.is_active == True).first()
        if not rule:
            raise ValueError(f"第{item.row_number}行规则已失效，请重新上传")
        score = calculate_rule_score(rule, item.role_rank)
        rank_match = re.search(r"\d+", item.role_rank or "")
        team_rank = max(1, int(rank_match.group(0))) if rank_match else 1
        application = ScoreApplication(
            student_id=student.id,
            rule_id=rule.id,
            project_name=item.project_name,
            project_level=item.project_level,
            team_rank=team_rank,
            team_total=max(1, team_rank),
            project_date=item.project_date,
            description=item.description,
            status=STATUS_DRAFT,
            system_calculated_score=score,
            final_score=score,
            remark=f"批次#{batch.id} Excel申报分:{item.declared_score:g} {item.evidence_ref}"[:200],
        )
        db.add(application)
        db.flush()
        for material in materials:
            db.add(EvidenceFile(
                application_id=application.id,
                file_name=material.original_name,
                file_path=material.stored_path,
                file_size=material.file_size,
                file_type=material.file_type,
            ))
        item.application_id = application.id
        application_ids.append(application.id)
    batch.status = BATCH_CONFIRMED
    batch.confirmed_at = datetime.now()
    log_operation(db, actor.id, "确认批量申报", student.id, "submission_batch", batch.id, f"生成 {len(application_ids)} 条草稿申请")
    return application_ids


def remove_batch_folder(path: Path):
    upload_root = Path(UPLOAD_DIR).resolve()
    resolved = path.resolve()
    if resolved != upload_root and upload_root in resolved.parents and path.exists() and path.is_dir():
        shutil.rmtree(path)
