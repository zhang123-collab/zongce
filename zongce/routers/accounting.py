from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io

from zongce.core import *
from zongce.accounting_service import (
    ASSESSMENT_YEAR,
    active_finalization,
    calculate_ranked_accounting,
    calculate_student_accounting,
    is_student_finalized,
    log_operation,
    sync_score_and_ranks,
)


router = APIRouter()


def _student_by_number(db: Session, student_id: str):
    return db.query(StudentProfile).filter(StudentProfile.student_id == student_id).first()


def _can_view_student(current_user: User, student: StudentProfile) -> bool:
    if current_user.role == ROLE_ADMIN:
        return True
    if current_user.role == ROLE_TEACHER:
        return bool(current_user.managed_class and current_user.managed_class == student.class_name)
    return current_user.id == student.user_id


@router.get("/api/accounting/me", tags=["🎓 1-学生端"], summary="学生：查看本人综合测评核算明细")
def accounting_me(
    current_user: User = Depends(require_role(ROLE_STUDENT)),
    db: Session = Depends(get_db),
):
    student = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    if not student:
        return ApiResponse.error(404, "学生档案不存在")
    ranked = calculate_ranked_accounting(db, [student])
    return ApiResponse.success(ranked[0] if ranked else calculate_student_accounting(db, student))


@router.get("/api/accounting/student/{student_id}", tags=["👤 4-通用"], summary="按权限查看学生核算明细")
def accounting_student_detail(
    student_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    student = _student_by_number(db, student_id)
    if not student:
        return ApiResponse.error(404, "学生不存在")
    if not _can_view_student(current_user, student):
        return ApiResponse.error(403, "无权查看该学生核算结果")
    ranked = calculate_ranked_accounting(db, [student])
    result = ranked[0] if ranked else calculate_student_accounting(db, student)
    logs = db.query(OperationLog).filter(OperationLog.student_id == student.id).order_by(OperationLog.id.desc()).limit(30).all()
    result["operationLogs"] = [
        {
            "id": item.id,
            "action": item.action,
            "detail": item.detail,
            "createdAt": item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else "",
        }
        for item in logs
    ]
    return ApiResponse.success(result)


@router.get("/api/accounting/list", tags=["👨‍🏫 2-审核端"], summary="教师/管理员：班级核算列表")
def accounting_list(
    class_name: str = "",
    keyword: str = "",
    current_user: User = Depends(require_role(ROLE_TEACHER, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    query = db.query(StudentProfile).join(StudentProfile.user)
    if current_user.role == ROLE_TEACHER:
        if not current_user.managed_class:
            return ApiResponse.success({"list": [], "managedClass": ""})
        query = query.filter(StudentProfile.class_name == current_user.managed_class)
    elif class_name:
        query = query.filter(StudentProfile.class_name == class_name)
    if keyword:
        query = query.filter((StudentProfile.student_id.contains(keyword)) | (User.real_name.contains(keyword)))
    selected = query.order_by(StudentProfile.class_name, StudentProfile.student_id).all()
    result = calculate_ranked_accounting(db, selected)
    return ApiResponse.success({"list": result, "managedClass": current_user.managed_class if current_user.role == ROLE_TEACHER else ""})


class BaseScoreReq(BaseModel):
    student_id: str
    moral_score: float
    academic_score: float


@router.post("/api/admin/accounting/base-score", tags=["🛡️ 3-管理端"], summary="管理员：保存学生基础分")
def accounting_base_score(
    req: BaseScoreReq,
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    student = _student_by_number(db, req.student_id)
    if not student:
        return ApiResponse.error(404, "学生不存在")
    if is_student_finalized(db, student.id):
        return ApiResponse.error(409, "该生核算结果已经终审，请先撤销终审")
    if not 0 <= req.moral_score <= 100 or not 0 <= req.academic_score <= 100:
        return ApiResponse.error(400, "思品分和学业成绩应在0-100之间")
    student.moral_score = req.moral_score
    student.academic_score = req.academic_score
    accounting = calculate_student_accounting(db, student, use_final_snapshot=False)
    sync_score_and_ranks(db, student, accounting)
    log_operation(db, current_user.id, "保存基础分", student.id, detail=f"思品 {req.moral_score:g}；学业 {req.academic_score:g}")
    db.commit()
    return ApiResponse.success(accounting, "基础分已保存并重新核算")


class DeductionCreateReq(BaseModel):
    student_id: str
    deduction_score: float
    reason: str
    scope: str = "综合测评总分"
    evidence_ref: str = ""
    rule_id: Optional[int] = None


@router.post("/api/admin/accounting/deduction", tags=["🛡️ 3-管理端"], summary="管理员：新增扣分认定")
def accounting_add_deduction(
    req: DeductionCreateReq,
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    student = _student_by_number(db, req.student_id)
    if not student:
        return ApiResponse.error(404, "学生不存在")
    if is_student_finalized(db, student.id):
        return ApiResponse.error(409, "该生核算结果已经终审，请先撤销终审")
    if not 0 < req.deduction_score <= 100:
        return ApiResponse.error(400, "扣分值必须大于0且不超过100")
    reason = (req.reason or "").strip()
    if not reason or len(reason) > 500:
        return ApiResponse.error(400, "扣分原因必填且不能超过500字")
    if req.scope not in ("思想品德", "学生工作", "综合测评总分"):
        return ApiResponse.error(400, "扣分作用范围无效")
    rule = db.query(ScoreRule).filter(ScoreRule.id == req.rule_id).first() if req.rule_id else None
    snapshot = f"{rule.category} / {rule.sub_category} / {rule.item_name}" if rule else "人工认定"
    deduction = ScoreDeduction(
        student_id=student.id,
        assessment_year=ASSESSMENT_YEAR,
        rule_id=rule.id if rule else None,
        rule_snapshot=snapshot,
        scope=req.scope,
        deduction_score=req.deduction_score,
        reason=reason,
        evidence_ref=(req.evidence_ref or "").strip()[:500],
        created_by=current_user.id,
    )
    db.add(deduction)
    db.flush()
    accounting = calculate_student_accounting(db, student, use_final_snapshot=False)
    sync_score_and_ranks(db, student, accounting)
    log_operation(db, current_user.id, "新增扣分", student.id, "deduction", deduction.id, f"-{req.deduction_score:g}；{reason}")
    db.commit()
    return ApiResponse.success(accounting, "扣分已记录并重新核算")


@router.delete("/api/admin/accounting/deduction/{deduction_id}", tags=["🛡️ 3-管理端"], summary="管理员：软删除扣分记录")
def accounting_delete_deduction(
    deduction_id: int,
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    deduction = db.query(ScoreDeduction).filter(ScoreDeduction.id == deduction_id, ScoreDeduction.is_active == True).first()
    if not deduction:
        return ApiResponse.error(404, "扣分记录不存在")
    if is_student_finalized(db, deduction.student_id, deduction.assessment_year):
        return ApiResponse.error(409, "该生核算结果已经终审，请先撤销终审")
    deduction.is_active = False
    deduction.deleted_by = current_user.id
    deduction.deleted_at = datetime.now()
    db.flush()
    student = db.query(StudentProfile).filter(StudentProfile.id == deduction.student_id).first()
    accounting = calculate_student_accounting(db, student, deduction.assessment_year, use_final_snapshot=False)
    sync_score_and_ranks(db, student, accounting)
    log_operation(db, current_user.id, "撤销扣分", student.id, "deduction", deduction.id, f"原扣分 {deduction.deduction_score:g}")
    db.commit()
    return ApiResponse.success(accounting, "扣分已撤销并重新核算")


@router.post("/api/admin/accounting/finalize/{student_id}", tags=["🛡️ 3-管理端"], summary="管理员：终审并锁定核算结果")
def accounting_finalize(
    student_id: str,
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    student = _student_by_number(db, student_id)
    if not student:
        return ApiResponse.error(404, "学生不存在")
    if is_student_finalized(db, student.id):
        return ApiResponse.error(409, "该生已经终审")
    pending = db.query(ScoreApplication).filter(
        ScoreApplication.student_id == student.id,
        ScoreApplication.status == STATUS_PENDING,
    ).count()
    if pending:
        return ApiResponse.error(400, f"仍有 {pending} 条申请正在审核，不能终审")
    accounting = calculate_student_accounting(db, student, use_final_snapshot=False)
    sync_score_and_ranks(db, student, accounting)
    finalization = db.query(ScoreFinalization).filter(
        ScoreFinalization.student_id == student.id,
        ScoreFinalization.assessment_year == ASSESSMENT_YEAR,
    ).first()
    if not finalization:
        finalization = ScoreFinalization(
            student_id=student.id,
            assessment_year=ASSESSMENT_YEAR,
            finalized_by=current_user.id,
        )
        db.add(finalization)
    finalization.is_finalized = True
    finalization.snapshot_json = json_lib.dumps(accounting, ensure_ascii=False)
    finalization.finalized_by = current_user.id
    finalization.finalized_at = datetime.now()
    finalization.reopened_by = None
    finalization.reopened_at = None
    finalization.reopen_reason = ""
    db.flush()
    log_operation(db, current_user.id, "终审核算结果", student.id, "finalization", finalization.id, f"总分 {accounting['totalScore']:g}")
    db.commit()
    accounting["isFinalized"] = True
    accounting["finalizedAt"] = finalization.finalized_at.strftime("%Y-%m-%d %H:%M:%S")
    return ApiResponse.success(accounting, "终审完成，成绩与申请已锁定")


class ReopenReq(BaseModel):
    student_id: str
    reason: str


@router.post("/api/admin/accounting/reopen", tags=["🛡️ 3-管理端"], summary="管理员：说明原因后撤销终审")
def accounting_reopen(
    req: ReopenReq,
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    student = _student_by_number(db, req.student_id)
    if not student:
        return ApiResponse.error(404, "学生不存在")
    reason = (req.reason or "").strip()
    if not reason or len(reason) > 500:
        return ApiResponse.error(400, "撤销终审原因必填且不能超过500字")
    finalization = active_finalization(db, student.id)
    if not finalization:
        return ApiResponse.error(400, "该生尚未终审")
    finalization.is_finalized = False
    finalization.reopened_by = current_user.id
    finalization.reopened_at = datetime.now()
    finalization.reopen_reason = reason
    log_operation(db, current_user.id, "撤销终审", student.id, "finalization", finalization.id, reason)
    db.commit()
    return ApiResponse.success(calculate_student_accounting(db, student, use_final_snapshot=False), "已撤销终审，可继续修正")


@router.get("/api/admin/accounting/overview", tags=["🛡️ 3-管理端"], summary="管理员：综合测评核算统计")
def accounting_overview(
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    students = db.query(StudentProfile).all()
    results = calculate_ranked_accounting(db, students)
    totals = [item["totalScore"] for item in results]
    buckets = [
        ("90及以上", 90, None),
        ("80-89.99", 80, 90),
        ("70-79.99", 70, 80),
        ("60-69.99", 60, 70),
        ("60以下", None, 60),
    ]
    distribution = [
        {
            "label": label,
            "count": sum(1 for score in totals if (low is None or score >= low) and (high is None or score < high)),
        }
        for label, low, high in buckets
    ]
    return ApiResponse.success({
        "studentCount": len(results),
        "finalizedCount": sum(1 for item in results if item["isFinalized"]),
        "pendingApplicationCount": db.query(ScoreApplication).filter(ScoreApplication.status == STATUS_PENDING).count(),
        "averageScore": _safe_average(totals),
        "highestScore": max(totals, default=0.0),
        "lowestScore": min(totals, default=0.0),
        "distribution": distribution,
    })


@router.post("/api/admin/accounting/recalculate-ranks", tags=["🛡️ 3-管理端"], summary="管理员：重新计算并保存班级/年级排名")
def recalculate_ranks(
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    results = calculate_ranked_accounting(db, persist=True)
    log_operation(db, current_user.id, "重新计算排名", detail=f"共处理{len(results)}名学生")
    db.commit()
    return ApiResponse.success({"count": len(results)}, "排名已重新计算")


@router.get("/api/admin/accounting/export", tags=["🛡️ 3-管理端"], summary="管理员：导出综合测评核算Excel")
def accounting_export(
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    students = db.query(StudentProfile).order_by(StudentProfile.class_name, StudentProfile.student_id).all()
    results = calculate_ranked_accounting(db, students)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "综合汇总"
    summary_sheet.append(["学号", "姓名", "班级", "思品", "学业", "学术创新", "学生工作", "扣分", "总分", "班级排名", "年级排名", "终审状态"])
    bonus_sheet = workbook.create_sheet("加分明细")
    bonus_sheet.append(["学号", "姓名", "类别", "子类", "聚合方式", "计入分"])
    deduction_sheet = workbook.create_sheet("扣分明细")
    deduction_sheet.append(["学号", "姓名", "规则快照", "作用范围", "扣分", "原因", "证据位置"])
    for result in results:
        summary_sheet.append([
            result["studentId"], result["studentName"], result["className"],
            result["moralScore"], result["academicScore"], result["innovationScore"],
            result["workScore"], result["deductionScore"], result["totalScore"],
            result.get("classRank"), result.get("gradeRank"),
            "已终审" if result["isFinalized"] else "核算中",
        ])
        for item in result["bonusDetails"]:
            bonus_sheet.append([
                result["studentId"], result["studentName"], item["category"], item["subCategory"],
                "取最高" if item["policy"] == POLICY_MAX else "累加", item["countedScore"],
            ])
        for item in result["deductions"]:
            deduction_sheet.append([
                result["studentId"], result["studentName"], item["ruleSnapshot"], item["scope"],
                item["score"], item["reason"], item["evidenceRef"],
            ])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
            sheet.column_dimensions[column[0].column_letter].width = max(width, 10)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"zongce_accounting_{ASSESSMENT_YEAR}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _safe_average(values: List[float]) -> float:
    return round(sum(values) / len(values), 4) if values else 0.0
