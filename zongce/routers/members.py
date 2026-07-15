from fastapi import APIRouter, Form
import io
import re

from zongce.core import *

router = APIRouter()


def validate_student_payload(data):
    username = (data.username or "").strip()
    student_id = (data.student_id or "").strip()
    real_name = (data.real_name or "").strip()
    class_name = (data.class_name or "").strip()
    major = (data.major or "").strip()
    grade = (data.grade or "").strip()
    if not username or len(username) > 50:
        return "登录账号不能为空且不能超过50个字符"
    if not re.fullmatch(r"\d{6,20}", student_id):
        return "学号应为6-20位数字"
    if not real_name or len(real_name) > 50:
        return "姓名不能为空且不能超过50个字符"
    if not class_name or len(class_name) > 50:
        return "班级不能为空且不能超过50个字符"
    if not major or len(major) > 100:
        return "专业不能为空且不能超过100个字符"
    if not re.fullmatch(r"\d{4}", grade):
        return "年级应为4位年份"
    if len(data.password or "") < 6:
        return "密码至少6位"
    if not 0 <= data.moral_score <= 100 or not 0 <= data.academic_score <= 100:
        return "思品分和学业成绩应在0-100之间"
    if data.email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", data.email.strip()):
        return "邮箱格式不正确"
    if data.phone and not re.fullmatch(r"[0-9+\-]{6,20}", data.phone.strip()):
        return "手机号格式不正确"
    return None

class LoginReq(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None


def _authenticate_user(username: str, password: str, db: Session):
    username = (username or "").strip()
    if not username or not password:
        return None, ApiResponse.error(400, "请输入用户名和密码")
    user = db.query(User).filter(User.username == username).first()
    if not user:
        return None, ApiResponse.error(400, "用户不存在，请检查用户名")
    if not user.is_active:
        return None, ApiResponse.error(403, "账号已停用，请联系管理员")
    now = datetime.now()
    if user.locked_until and user.locked_until > now:
        remaining = max(1, int((user.locked_until - now).total_seconds() // 60) + 1)
        return None, ApiResponse.error(423, f"密码错误次数过多，请{remaining}分钟后重试")
    if user.locked_until and user.locked_until <= now:
        user.locked_until = None
        user.failed_login_count = 0
    if not verify_password(password, user.password):
        user.failed_login_count = (user.failed_login_count or 0) + 1
        if user.failed_login_count >= MAX_LOGIN_ATTEMPTS:
            user.locked_until = now + timedelta(minutes=LOGIN_LOCK_MINUTES)
            message = f"密码错误次数过多，账号已锁定{LOGIN_LOCK_MINUTES}分钟"
        else:
            message = f"密码错误，还可尝试{MAX_LOGIN_ATTEMPTS - user.failed_login_count}次"
        db.commit()
        return None, ApiResponse.error(400, message)
    user.failed_login_count = 0
    user.locked_until = None
    user.last_login_at = now
    db.commit()
    return user, None


def _do_login(username: str, password: str, db: Session):
    user, error = _authenticate_user(username, password, db)
    if error:
        return error
    access_token = create_access_token(data={"user_id": user.id, "role": user.role})
    role_name = ["学生", "审核老师", "系统管理员"][user.role]
    data = {
        "token": access_token,
        "access_token": access_token,
        "token_type": "bearer",
        "expires_in": ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        "userId": user.id,
        "user_id": user.id,
        "username": user.username,
        "realName": user.real_name,
        "real_name": user.real_name,
        "role": user.role,
        "roleName": role_name,
        "role_name": role_name,
        "isAdmin": user.role == ROLE_ADMIN,
        "isTeacher": user.role == ROLE_TEACHER,
        "isStudent": user.role == ROLE_STUDENT,
        "is_admin": user.role == ROLE_ADMIN,
        "is_teacher": user.role == ROLE_TEACHER,
        "is_student": user.role == ROLE_STUDENT,
        "email": user.email,
        "phone": user.phone,
        "permissions": self_permissions(user.role)
    }
    if user.role == ROLE_STUDENT:
        sp = db.query(StudentProfile).filter(StudentProfile.user_id == user.id).first()
        if sp:
            data["studentId"] = sp.student_id
            data["className"] = sp.class_name
            data["major"] = sp.major
            data["grade"] = sp.grade
    return ApiResponse.success(data, "登录成功")

def self_permissions(role: int) -> List[str]:
    base = ["profile:read", "profile:update", "application:list", "application:detail", "file:upload"]
    if role == ROLE_STUDENT:
        return base + ["application:create", "application:submit", "application:withdraw", "application:delete", "score:view", "objection:submit"]
    if role == ROLE_TEACHER:
        return base + ["audit:list", "audit:detail", "audit:pass", "audit:reject", "audit:modify", "audit:history"]
    if role == ROLE_ADMIN:
        return base + ["audit:*", "student:*", "rule:*", "score:*", "export:*", "announcement:*", "objection:*", "system:*"]
    return base

@router.post("/api/oauth/token", include_in_schema=True, summary="Swagger OAuth2 专用登录", tags=["🔓 0-登录与公共"])
async def oauth_token(
    username: str = Form(...),
    password: str = Form(...),
    grant_type: Optional[str] = Form(default=None),
    scope: str = Form(default=""),
    client_id: Optional[str] = Form(default=None),
    client_secret: Optional[str] = Form(default=None),
    db: Session = Depends(get_db)
):
    from fastapi.responses import JSONResponse as JR
    if not username or not password:
        return JR(status_code=400, content={"error": "invalid_request", "error_description": "请输入用户名和密码"})
    user, error = _authenticate_user(username, password, db)
    if error:
        return JR(status_code=401, content={"error": "invalid_grant", "error_description": error["message"]})
    access_token = create_access_token(data={"user_id": user.id, "role": user.role})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "expires_in": ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        "scope": scope
    }

@router.post("/api/login", tags=["🔓 0-登录与公共"], summary="登录账号")
async def login(
    request: Request,
    username: Optional[str] = Form(default=None),
    password: Optional[str] = Form(default=None),
    db: Session = Depends(get_db)
):
    try:
        body = await request.json()
    except Exception:
        body = None
    if not username and body and isinstance(body, dict):
        username = body.get("username") or body.get("userName") or body.get("account")
    if not password and body and isinstance(body, dict):
        password = body.get("password") or body.get("passWord") or body.get("pwd")
    return _do_login(username or "", password or "", db)

@router.post("/api/logout", tags=["👤 4-通用"], summary="退出登录")
def logout(current_user: User = Depends(get_current_user)):
    return ApiResponse.success(message="退出登录成功")

@router.get("/api/user/info", tags=["👤 4-通用"], summary="获取当前登录用户信息")
def get_user_info(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    role_name = ["学生", "审核老师", "系统管理员"][current_user.role]
    data = {
        "userId": current_user.id,
        "user_id": current_user.id,
        "username": current_user.username,
        "realName": current_user.real_name,
        "real_name": current_user.real_name,
        "role": current_user.role,
        "roleName": role_name,
        "role_name": role_name,
        "isAdmin": current_user.role == ROLE_ADMIN,
        "isTeacher": current_user.role == ROLE_TEACHER,
        "isStudent": current_user.role == ROLE_STUDENT,
        "is_admin": current_user.role == ROLE_ADMIN,
        "is_teacher": current_user.role == ROLE_TEACHER,
        "is_student": current_user.role == ROLE_STUDENT,
        "email": current_user.email,
        "phone": current_user.phone,
        "permissions": self_permissions(current_user.role)
    }
    if current_user.role == ROLE_STUDENT:
        sp = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
        if sp:
            data["studentId"] = sp.student_id
            data["className"] = sp.class_name
            data["major"] = sp.major
            data["grade"] = sp.grade
    return ApiResponse.success(data)

@router.get("/api/student/profile", tags=["🎓 1-学生端"], summary="获取学生自己的基本信息")
def get_student_profile(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sp = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    if not sp:
        return ApiResponse.error(404, "学生信息不存在")
    return ApiResponse.success({
        "id": sp.id,
        "userId": sp.user_id,
        "username": current_user.username,
        "studentId": sp.student_id,
        "className": sp.class_name,
        "major": sp.major,
        "grade": sp.grade,
        "moralScore": sp.moral_score,
        "academicScore": sp.academic_score,
        "realName": current_user.real_name,
        "email": current_user.email,
        "phone": current_user.phone,
        "age": current_user.age,
    })

class StudentProfileUpdate(BaseModel):
    password: Optional[str] = None
    new_password: Optional[str] = None
    age: Optional[int] = None
    email: Optional[str] = None
    phone: Optional[str] = None

@router.put("/api/student/profile", tags=["🎓 1-学生端"], summary="学生更新：密码/年龄/邮箱/电话")
def update_student_profile(req: StudentProfileUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sp = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    if not sp:
        return ApiResponse.error(404, "学生信息不存在")
    if req.new_password:
        if len(req.new_password) < 6:
            return ApiResponse.error(400, "新密码至少6位")
        if not req.password or not verify_password(req.password, current_user.password):
            return ApiResponse.error(400, "旧密码不正确")
        current_user.password = get_password_hash(req.new_password)
    if req.email is not None:
        current_user.email = req.email
    if req.phone is not None:
        current_user.phone = req.phone
    if req.age is not None:
        if req.age < 15 or req.age > 100:
            return ApiResponse.error(400, "年龄应在15-100之间")
        current_user.age = req.age
    db.commit()
    return ApiResponse.success(message="更新成功")

@router.get("/api/admin/student/list", tags=["🛡️ 3-管理端"], summary="管理员：学生列表（分页+筛选）")
def admin_student_list(
    page: int = 1, page_size: int = 20, keyword: str = "",
    class_name: str = "", major: str = "",
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db)
):
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    query = db.query(StudentProfile).join(User)
    if keyword:
        query = query.filter((StudentProfile.student_id.contains(keyword)) | (User.real_name.contains(keyword)))
    if class_name:
        query = query.filter(StudentProfile.class_name == class_name)
    if major:
        query = query.filter(StudentProfile.major == major)
    total = query.count()
    items = query.order_by(StudentProfile.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    rows = []
    for sp in items:
        rows.append({
            "id": sp.id,
            "userId": sp.user_id,
            "studentId": sp.student_id,
            "className": sp.class_name,
            "major": sp.major,
            "grade": sp.grade,
            "moralScore": sp.moral_score,
            "academicScore": sp.academic_score,
            "realName": sp.user.real_name if sp.user else "",
            "username": sp.user.username if sp.user else ""
            ,"isActive": bool(sp.user.is_active) if sp.user else False
        })
    return ApiResponse.success({"total": total, "page": page, "pageSize": page_size, "list": rows})

class AdminStudentCreate(BaseModel):
    username: str
    password: str = "123456"
    real_name: str
    student_id: str
    class_name: str
    major: str
    grade: str = "2023"
    moral_score: float = 0.0
    academic_score: float = 0.0
    email: str = ""
    phone: str = ""

@router.post("/api/admin/student/create", tags=["🛡️ 3-管理端"], summary="管理员：新增学生账号和档案")
def admin_student_create(req: AdminStudentCreate, current_user: User = Depends(require_role(ROLE_ADMIN)), db: Session = Depends(get_db)):
    validation_error = validate_student_payload(req)
    if validation_error:
        return ApiResponse.error(400, validation_error)
    req.username = req.username.strip()
    req.student_id = req.student_id.strip()
    req.real_name = req.real_name.strip()
    req.class_name = req.class_name.strip()
    req.major = req.major.strip()
    if db.query(User).filter(User.username == req.username).first():
        return ApiResponse.error(409, "用户名已存在")
    if db.query(StudentProfile).filter(StudentProfile.student_id == req.student_id).first():
        return ApiResponse.error(409, "该学号已存在，请勿重复录入")
    user = User(
        username=req.username,
        password=get_password_hash(req.password),
        role=ROLE_STUDENT,
        real_name=req.real_name,
        email=req.email,
        phone=req.phone
    )
    db.add(user)
    db.flush()
    sp = StudentProfile(
        user_id=user.id,
        student_id=req.student_id,
        class_name=req.class_name,
        major=req.major,
        grade=req.grade,
        moral_score=req.moral_score,
        academic_score=req.academic_score
    )
    db.add(sp)
    db.commit()
    return ApiResponse.success(message="学生创建成功")


@router.post("/api/admin/student/import", tags=["🛡️ 3-管理端"], summary="管理员：批量导入学生 Excel")
async def admin_student_import(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        return ApiResponse.error(400, "仅支持 .xlsx 文件")
    content = await file.read()
    if not content or len(content) > 5 * 1024 * 1024:
        return ApiResponse.error(400, "Excel不能为空且不能超过5MB")
    from openpyxl import load_workbook
    try:
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
    except Exception:
        return ApiResponse.error(400, "Excel文件损坏或格式不正确")
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows, [])]
    required = ["学号", "姓名", "班级", "专业", "年级", "思品分", "学业成绩"]
    missing = [name for name in required if name not in headers]
    if missing:
        return ApiResponse.error(400, "缺少列：" + "、".join(missing))
    index = {name: headers.index(name) for name in headers}
    success = 0
    failures = []
    for row_number, row in enumerate(rows, start=2):
        if row_number > 1001:
            failures.append({"row": row_number, "reason": "单次最多导入1000名学生"})
            break
        if not any(value is not None and str(value).strip() for value in row):
            continue
        def value(name, default=""):
            position = index.get(name)
            return row[position] if position is not None and position < len(row) and row[position] is not None else default
        try:
            request = AdminStudentCreate(
                username=str(value("登录账号", value("学号"))).strip(),
                password=str(value("初始密码", "123456")),
                real_name=str(value("姓名")).strip(),
                student_id=str(value("学号")).strip().removesuffix(".0"),
                class_name=str(value("班级")).strip(),
                major=str(value("专业")).strip(),
                grade=str(value("年级")).strip().removesuffix(".0"),
                moral_score=float(value("思品分", 0)),
                academic_score=float(value("学业成绩", 0)),
                email=str(value("邮箱", "")).strip(),
                phone=str(value("电话", "")).strip().removesuffix(".0"),
            )
        except Exception:
            failures.append({"row": row_number, "reason": "字段类型不正确"})
            continue
        error = validate_student_payload(request)
        if error:
            failures.append({"row": row_number, "studentId": request.student_id, "reason": error})
            continue
        if db.query(User).filter(User.username == request.username).first() or db.query(StudentProfile).filter(StudentProfile.student_id == request.student_id).first():
            failures.append({"row": row_number, "studentId": request.student_id, "reason": "账号或学号已存在"})
            continue
        user = User(
            username=request.username, password=get_password_hash(request.password),
            role=ROLE_STUDENT, real_name=request.real_name,
            email=request.email, phone=request.phone,
        )
        db.add(user)
        db.flush()
        db.add(StudentProfile(
            user_id=user.id, student_id=request.student_id,
            class_name=request.class_name, major=request.major, grade=request.grade,
            moral_score=request.moral_score, academic_score=request.academic_score,
        ))
        success += 1
    db.commit()
    return ApiResponse.success({"success": success, "failed": len(failures), "failures": failures}, "导入完成")

@router.delete("/api/admin/student/delete/{student_id}", tags=["🛡️ 3-管理端"], summary="管理员：删除学生（级联删申请/材料）")
def admin_student_delete(student_id: int, current_user: User = Depends(require_role(ROLE_ADMIN)), db: Session = Depends(get_db)):
    sp = db.query(StudentProfile).filter(StudentProfile.id == student_id).first()
    if not sp:
        return ApiResponse.error(404, "学生不存在")
    user = sp.user
    application_count = db.query(ScoreApplication).filter(ScoreApplication.student_id == sp.id).count()
    if application_count:
        user.is_active = False
        db.commit()
        return ApiResponse.success(message=f"学生存在{application_count}条业务记录，账号已停用并保留历史数据")
    db.delete(sp)
    if user:
        db.delete(user)
    db.commit()
    return ApiResponse.success(message="删除成功")

class UserProfileUpdate(BaseModel):
    password: Optional[str] = None
    new_password: Optional[str] = None
    age: Optional[int] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    managed_class: Optional[str] = None
    real_name: Optional[str] = None

@router.get("/api/user/profile", tags=["👤 4-通用"], summary="获取当前用户 profile（教师/管理员/学生通用）")
def get_user_profile(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    data = {
        "id": current_user.id,
        "username": current_user.username,
        "role": current_user.role,
        "roleName": ["学生", "教师", "管理员"][current_user.role],
        "realName": current_user.real_name,
        "email": current_user.email,
        "phone": current_user.phone,
        "age": current_user.age,
        "managedClass": current_user.managed_class,
    }
    if current_user.role == ROLE_STUDENT:
        sp = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
        if sp:
            data.update({
                "studentId": sp.student_id,
                "className": sp.class_name,
                "major": sp.major,
                "grade": sp.grade,
                "moralScore": sp.moral_score,
                "academicScore": sp.academic_score,
            })
    return ApiResponse.success(data)

@router.put("/api/user/profile", tags=["👤 4-通用"], summary="更新当前用户 profile")
def update_user_profile(req: UserProfileUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if req.new_password:
        if len(req.new_password) < 6:
            return ApiResponse.error(400, "新密码至少6位")
        if not req.password or not verify_password(req.password, current_user.password):
            return ApiResponse.error(400, "旧密码不正确")
        current_user.password = get_password_hash(req.new_password)
    if req.real_name is not None:
        current_user.real_name = req.real_name
    if req.email is not None:
        current_user.email = req.email
    if req.phone is not None:
        current_user.phone = req.phone
    if req.age is not None:
        if req.age < 15 or req.age > 100:
            return ApiResponse.error(400, "年龄应在15-100之间")
        current_user.age = req.age
    db.commit()
    return ApiResponse.success(message="更新成功")
