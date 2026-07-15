from pathlib import Path
from uuid import uuid4
import os

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io

from zongce.core import *
from zongce.accounting_service import ASSESSMENT_YEAR, is_student_finalized, log_operation
from zongce.batch_service import *


router = APIRouter()


def _batch_data(db: Session, batch: SubmissionBatch):
    items = db.query(SubmissionBatchItem).filter(SubmissionBatchItem.batch_id == batch.id).order_by(SubmissionBatchItem.row_number).all()
    materials = db.query(SubmissionBatchMaterial).filter(
        SubmissionBatchMaterial.batch_id == batch.id,
        SubmissionBatchMaterial.is_active == True,
    ).order_by(SubmissionBatchMaterial.id).all()
    latest_ai_job = db.query(AiAnalysisJob).filter(
        AiAnalysisJob.batch_id == batch.id,
    ).order_by(AiAnalysisJob.id.desc()).first()
    latest_suggestions = db.query(AiItemSuggestion).filter(
        AiItemSuggestion.job_id == latest_ai_job.id,
    ).order_by(AiItemSuggestion.batch_item_id).all() if latest_ai_job else []
    return {
        "id": batch.id,
        "assessmentYear": batch.assessment_year,
        "status": batch.status,
        "version": batch.version,
        "excelName": batch.excel_name,
        "itemCount": batch.item_count,
        "validCount": batch.valid_count,
        "errorCount": batch.error_count,
        "createdAt": batch.created_at.strftime("%Y-%m-%d %H:%M:%S") if batch.created_at else "",
        "latestAiJob": {
            "id": latest_ai_job.id,
            "status": latest_ai_job.status,
            "completedCount": latest_ai_job.completed_count,
            "errorMessage": latest_ai_job.error_message,
            "suggestions": [{
                "itemId": suggestion.batch_item_id,
                "verificationStatus": suggestion.verification_status,
                "suggestedScore": suggestion.suggested_score,
                "reason": suggestion.reason,
            } for suggestion in latest_suggestions],
        } if latest_ai_job else None,
        "items": [{
            "id": item.id,
            "rowNumber": item.row_number,
            "ruleId": item.rule_id,
            "projectName": item.project_name,
            "projectLevel": item.project_level,
            "roleRank": item.role_rank,
            "declaredScore": item.declared_score,
            "projectDate": item.project_date,
            "evidenceRef": item.evidence_ref,
            "status": item.status,
            "errorMessage": item.error_message,
            "applicationId": item.application_id,
        } for item in items],
        "materials": [{
            "id": material.id,
            "filename": material.original_name,
            "size": material.file_size,
            "url": f"/api/batch/material/{material.id}",
        } for material in materials],
    }


@router.post("/api/batch/upload", tags=["🎓 1-学生端"], summary="学生：上传一份申报Excel和多份证明材料")
async def batch_upload(
    excel: UploadFile = File(...),
    materials: List[UploadFile] = File(default=[]),
    current_user: User = Depends(require_role(ROLE_STUDENT)),
    db: Session = Depends(get_db),
):
    student = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    if not student:
        return ApiResponse.error(404, "学生档案不存在")
    if is_student_finalized(db, student.id):
        return ApiResponse.error(409, "综合测评结果已经终审，不能上传新批次")
    if not materials:
        return ApiResponse.error(400, "请至少上传一份证明材料")
    if len(materials) > BATCH_MAX_MATERIALS:
        return ApiResponse.error(400, f"单次最多上传{BATCH_MAX_MATERIALS}份证明材料")
    folder = Path(UPLOAD_DIR) / "batches" / f"u{student.id}_{uuid4().hex}"
    try:
        excel_content = await excel.read()
        excel_name = validate_excel_content(excel.filename or "", excel_content)
        material_payloads = []
        total_size = 0
        for material in materials:
            content = await material.read()
            name, ext, media_type = validate_material_content(material.filename or "", content)
            total_size += len(content)
            if total_size > TOTAL_MATERIAL_MAX_BYTES:
                raise ValueError("证明材料总大小不能超过30MB")
            material_payloads.append((name, ext, media_type, content))
        folder.mkdir(parents=True, exist_ok=False)
        excel_path = folder / f"{uuid4().hex}.xlsx"
        excel_path.write_bytes(excel_content)
        material_records = []
        for name, ext, media_type, content in material_payloads:
            path = folder / f"{uuid4().hex}{ext}"
            path.write_bytes(content)
            material_records.append((name, path, media_type, len(content)))
        parsed_items = parse_batch_excel(db, str(excel_path), student)
        valid_count = sum(1 for item in parsed_items if item["status"] == "valid")
        error_count = len(parsed_items) - valid_count
        batch = SubmissionBatch(
            student_id=student.id,
            assessment_year=ASSESSMENT_YEAR,
            status=BATCH_NEEDS_CORRECTION if error_count else BATCH_PARSED,
            excel_name=excel_name,
            excel_path=str(excel_path),
            item_count=len(parsed_items),
            valid_count=valid_count,
            error_count=error_count,
        )
        db.add(batch)
        db.flush()
        for item in parsed_items:
            db.add(SubmissionBatchItem(batch_id=batch.id, **{key: item[key] for key in (
                "row_number", "rule_id", "project_name", "project_level", "role_rank",
                "declared_score", "project_date", "description", "evidence_ref", "status", "error_message",
            )}))
        for name, path, media_type, size in material_records:
            db.add(SubmissionBatchMaterial(
                batch_id=batch.id,
                original_name=name,
                stored_path=str(path),
                file_size=size,
                file_type=media_type,
            ))
        log_operation(db, current_user.id, "上传批量申报", student.id, "submission_batch", batch.id, f"{len(parsed_items)}条明细；{len(material_records)}份材料")
        db.commit()
        db.refresh(batch)
        return ApiResponse.success(_batch_data(db, batch), "解析完成" if not error_count else "解析完成，请先修正错误行")
    except ValueError as exc:
        db.rollback()
        remove_batch_folder(folder)
        return ApiResponse.error(400, str(exc))
    except Exception:
        db.rollback()
        remove_batch_folder(folder)
        raise


@router.get("/api/batch/my", tags=["🎓 1-学生端"], summary="学生：查看自己的批量申报记录")
def batch_my(
    current_user: User = Depends(require_role(ROLE_STUDENT)),
    db: Session = Depends(get_db),
):
    student = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    if not student:
        return ApiResponse.error(404, "学生档案不存在")
    batches = db.query(SubmissionBatch).filter(SubmissionBatch.student_id == student.id).order_by(SubmissionBatch.id.desc()).all()
    return ApiResponse.success({"list": [_batch_data(db, batch) for batch in batches]})


@router.post("/api/batch/{batch_id}/confirm", tags=["🎓 1-学生端"], summary="学生：确认批次并生成草稿申请")
def batch_confirm(
    batch_id: int,
    current_user: User = Depends(require_role(ROLE_STUDENT)),
    db: Session = Depends(get_db),
):
    student = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
    batch = db.query(SubmissionBatch).filter(SubmissionBatch.id == batch_id).first()
    if not student or not batch:
        return ApiResponse.error(404, "批次不存在")
    try:
        application_ids = confirm_batch(db, batch, student, current_user)
        db.commit()
        return ApiResponse.success({"applicationIds": application_ids}, f"已生成{len(application_ids)}条草稿申请")
    except ValueError as exc:
        db.rollback()
        return ApiResponse.error(400, str(exc))


@router.get("/api/batch/material/{material_id}", tags=["👤 4-通用"], summary="按权限查看批次证明材料")
def batch_material(
    material_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    material = db.query(SubmissionBatchMaterial).filter(SubmissionBatchMaterial.id == material_id).first()
    batch = db.query(SubmissionBatch).filter(SubmissionBatch.id == material.batch_id).first() if material else None
    student = db.query(StudentProfile).filter(StudentProfile.id == batch.student_id).first() if batch else None
    allowed = bool(student and (
        current_user.role == ROLE_ADMIN
        or (current_user.role == ROLE_TEACHER and current_user.managed_class and current_user.managed_class == student.class_name)
        or current_user.id == student.user_id
    ))
    if not material or not material.is_active or not allowed:
        return ApiResponse.error(403, "无权查看该材料")
    if not os.path.isfile(material.stored_path):
        return ApiResponse.error(404, "材料文件已丢失")
    return FileResponse(material.stored_path, media_type=material.file_type, filename=material.original_name, content_disposition_type="inline")


@router.get("/api/batch/template", tags=["🎓 1-学生端"], summary="下载批量申报Excel模板")
def batch_template(current_user: User = Depends(require_role(ROLE_STUDENT))):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "学生汇总"
    summary.append(["姓名", "学号", "备注"])
    summary.append([current_user.real_name, "请填写本人学号", "学号和姓名必须与当前登录账号一致"])
    details = workbook.create_sheet("申报明细")
    details.append(["申报明细（学生逐项填写，系统优先读取本页）"] * 16)
    details.append([
        "明细序号", "模块", "汇总类别", "项目/活动名称", "规则等级/获奖结果", "本人身份/排名",
        "学生申报分", "证据编号", "证据类型", "公用材料页码", "个人材料页码", "证据说明/关键词",
        "发生/获奖/任职时间", "姓名核验", "时间核验", "备注",
    ])
    for index in range(1, 11):
        details.append([index, "", "", "", "", "", "", f"E{index:02d}", "", "", "", "", "", "待核验", "待核验", ""])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")
        sheet.freeze_panes = "A3" if sheet.title == "申报明细" else "A2"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="zongce_submission_template.xlsx"'},
    )
